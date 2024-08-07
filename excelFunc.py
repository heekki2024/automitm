

from tkinter import Tk, filedialog, messagebox
import tkinter as tk
import openpyxl
import configparser
import os
import sys
import pdb  



CONFIG_FILE = 'config.ini'

def load_config():
    config = configparser.ConfigParser()
    
    if os.path.exists(CONFIG_FILE):
        config.read(CONFIG_FILE)


        return config

    else:
        config = first_set_paths_gui()


        return config

def initialize_paths(config, root):
    if (os.path.exists(config['PATHS']['input_path']) 
        and os.path.exists(config['PATHS']['output_path'])
        and os.path.exists(config['PATHS']['base_path'])
        and os.path.exists(config['PATHS']['network_security_config_path'])
        and os.path.exists(config['PATHS']['network_security_config_with_r_path'])):
        excel_input_path = config['PATHS']['input_path']
        excel_output_path = config['PATHS']['output_path']
        base_path = config['PATHS']['base_path']
        network_security_config_path = config['PATHS']['network_security_config_path']    
        network_security_config_with_r_path = config['PATHS']['network_security_config_with_r_path']

        print(f"사용할 입력 파일 경로: {excel_input_path}")
        print(f"사용할 출력 파일 경로: {excel_output_path}")
        print(f"사용할 Base 폴더 경로: {base_path}")
        print(f"사용할 network_security_config 경로: {network_security_config_path}")
        print(f"사용할 network_security_config_with_r_path 경로: {network_security_config_with_r_path}")
    
    else:
        config = set_paths_gui(root)
        excel_input_path = config['PATHS']['input_path']
        excel_output_path = config['PATHS']['output_path']
        base_path = config['PATHS']['base_path']
        network_security_config_path = config['PATHS']['network_security_config_path']    
        network_security_config_with_r_path = config['PATHS']['network_security_config_with_r_path']

        print(f"사용할 입력 파일 경로: {excel_input_path}")
        print(f"사용할 출력 파일 경로: {excel_output_path}")
        print(f"사용할 Base 폴더 경로: {base_path}")
        print(f"사용할 network_security_config 경로: {network_security_config_path}")
        print(f"사용할 network_security_config_with_r_path 경로: {network_security_config_with_r_path}")

    return excel_input_path, excel_output_path, base_path, network_security_config_path, network_security_config_with_r_path

def save_config(config):
    with open(CONFIG_FILE, 'w') as configfile:
        config.write(configfile)

def select_file_path():
    root = Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename(
        title="파일을 선택하세요",
        #튜플타입
        filetypes=[("모든 파일", "*.*")]
    )
    root.destroy()
    return file_path

def select_folder_path():
    root = tk.Tk()
    root.withdraw()  # 메인 윈도우를 숨깁니다
    folder_path = filedialog.askdirectory(
        title="폴더를 선택하세요"
    )
    root.destroy()  # 메인 윈도우를 종료합니다
    return folder_path

def center_window(root, width, height):
    # 화면 너비와 높이 가져오기
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()

    # 창의 위치 계산
    x = (screen_width - width) // 2
    y = (screen_height - height) // 2

    # 창의 위치와 크기 설정
    root.geometry(f'{width}x{height}+{x}+{y}')

def first_set_paths_gui():
    root = tk.Tk()
    root.title("첫 파일 경로 설정")
    # root.attributes('-topmost', True)
    center_window(root, 700, 230)

    def on_closing():
        root.destroy()
        sys.exit()
        
    root.protocol("WM_DELETE_WINDOW", on_closing)


    config = configparser.ConfigParser()
    config.read(CONFIG_FILE)

    config['PATHS'] = {
        'input_path': '',
        'output_path': '',
        'base_path': '',
        'network_security_config_path': '',
        'network_security_config_with_r_path': ''
    }

    input_var = tk.StringVar(value=config['PATHS']['input_path'])

    output_var = tk.StringVar(value=config['PATHS']['output_path'])

    base_var = tk.StringVar(value=config['PATHS']['base_path'])

    network_security_config_path_var = tk.StringVar(value=config['PATHS']['network_security_config_path'])

    network_security_config_with_r_path_var = tk.StringVar(value=config['PATHS']['network_security_config_with_r_path'])
   
    def update_input_path():
        new_path = select_file_path()
        if new_path:
            input_var.set(new_path)
            print(f"Input path updated to: {new_path}")  # 디버깅 출력
            print(f"input_var value: {input_var.get()}")  # 추가 디버깅 출력

    def update_output_path():
        new_path = select_file_path()
        if new_path:
            output_var.set(new_path)

    def update_base_path():
        new_path = select_folder_path()
        if new_path:
            base_var.set(new_path)

    def update_network_security_config_path_path():
        new_path = select_file_path()
        if new_path:
            network_security_config_path_var.set(new_path)

    def update_network_security_config_with_r_path():
        new_path = select_file_path()
        if new_path:
            network_security_config_with_r_path_var.set(new_path)

    tk.Label(root, text="Excel 입력 파일 경로 :").grid(row=0, column=0, padx=10, pady=5)
    tk.Entry(root, textvariable=input_var, width=50).grid(row=0, column=1, padx=10, pady=5)
    tk.Button(root, text="변경", command=update_input_path).grid(row=0, column=2, padx=10, pady=5)

    tk.Label(root, text="Excel 출력 파일 경로 :").grid(row=1, column=0, padx=10, pady=5)
    tk.Entry(root, textvariable=output_var, width=50).grid(row=1, column=1, padx=10, pady=5)
    tk.Button(root, text="변경", command=update_output_path).grid(row=1, column=2, padx=10, pady=5)

    tk.Label(root, text="Base 폴더 경로 :").grid(row=2, column=0, padx=10, pady=5)
    tk.Entry(root, textvariable=base_var, width=50).grid(row=2, column=1, padx=10, pady=5)
    tk.Button(root, text="변경", command=update_base_path).grid(row=2, column=2, padx=10, pady=5)

    tk.Label(root, text="network_security_config 파일 경로 :").grid(row=3, column=0, padx=10, pady=5)
    tk.Entry(root, textvariable=network_security_config_path_var, width=50).grid(row=3, column=1, padx=10, pady=5)
    tk.Button(root, text="변경", command=update_network_security_config_path_path).grid(row=3, column=2, padx=10, pady=5)

    tk.Label(root, text="network_security_config_with_r 파일 경로 :").grid(row=4, column=0, padx=10, pady=5)
    tk.Entry(root, textvariable=network_security_config_with_r_path_var, width=50).grid(row=4, column=1, padx=10, pady=5)
    tk.Button(root, text="변경", command=update_network_security_config_with_r_path).grid(row=4, column=2, padx=10, pady=5)

    def save_and_close():
        if not input_var.get() or not output_var.get() or not base_var.get() or not network_security_config_path_var.get() or not network_security_config_with_r_path_var.get():
            messagebox.showwarning("경고", "모든 경로를 설정해야 합니다.")
            return  # 경로가 지정되지 않은 경우 함수를 종료하고 계속 실행     
           
        config['PATHS']['input_path'] = input_var.get()
        config['PATHS']['output_path'] = output_var.get()
        config['PATHS']['base_path'] = base_var.get()
        config['PATHS']['network_security_config_path'] = network_security_config_path_var.get()      
        config['PATHS']['network_security_config_with_r_path'] = network_security_config_with_r_path_var.get()        

        save_config(config)
        root.destroy()



    tk.Button(root, text=" 저장 ", command=save_and_close).grid(row=5, column=0, columnspan=3, pady=10)

    root.mainloop()

    # return config['PATHS']['input_path'], config['PATHS']['output_path']
    return config


def set_paths_gui(root):
    set_path_window = tk.Toplevel(root)
    set_path_window.title("파일 경로 설정")
    # set_path_window.attributes('-topmost', True)
    center_window(set_path_window, 700, 230)

    def on_closing():
        set_path_window.destroy()
        # sys.exit()
    set_path_window.protocol("WM_DELETE_WINDOW", on_closing)


    config = configparser.ConfigParser()

    # pdb.set_trace()  # Add a breakpoint here to inspect the config object

    # (Pdb) config.sections()  # config의 섹션 목록을 출력
    # (Pdb) config.items('PATHS')  # 'PATHS' 섹션의 모든 키-값 쌍을 출력
    # (Pdb) config['PATHS']['input_path']  # 'input_path' 값 출력
    # (Pdb) config['PATHS']['output_path']  # 'output_path' 값 출력

    if os.path.exists(CONFIG_FILE):
        config.read(CONFIG_FILE)
        if not os.path.exists(config['PATHS']['input_path']):
            config['PATHS'] = {'input_path': ''}
            input_var = tk.StringVar(value=config['PATHS']['input_path'])
        else:
            input_var = tk.StringVar(value=config['PATHS']['input_path'])
            

        if not os.path.exists(config['PATHS']['output_path']):
            config['PATHS'] = {'output_path': ''}
            output_var = tk.StringVar(value=config['PATHS']['output_path'])

        else:
            output_var = tk.StringVar(value=config['PATHS']['output_path'])
        

        if not os.path.exists(config['PATHS']['base_path']):
            config['PATHS'] = {'base_path': ''}
            base_var = tk.StringVar(value=config['PATHS']['base_path'])

        else:
            base_var = tk.StringVar(value=config['PATHS']['base_path'])


        if not os.path.exists(config['PATHS']['network_security_config_path']):
            config['PATHS'] = {'network_security_config_path': ''}
            network_security_config_path_var = tk.StringVar(value=config['PATHS']['network_security_config_path'])
        else:
            network_security_config_path_var = tk.StringVar(value=config['PATHS']['network_security_config_path'])


        if not os.path.exists(config['PATHS']['network_security_config_with_r_path']):
            config['PATHS'] = {'network_security_config_with_r_path': ''}
            network_security_config_with_r_path_var = tk.StringVar(value=config['PATHS']['network_security_config_with_r_path'])
   
        else:
            network_security_config_with_r_path_var = tk.StringVar(value=config['PATHS']['network_security_config_with_r_path'])
        
    else:
        # config 파일이 존재하지 않을 경우 한번에 config설정. 값은 없음
        config['PATHS'] = {'input_path': '', 'output_path': '', 'base_path': '', 'network_security_config_path': '', 'network_security_config_with_r_path': ''}

    # input_var = tk.StringVar(value=config['PATHS']['input_path'])
    # output_var = tk.StringVar(value=config['PATHS']['output_path'])
    # base_var = tk.StringVar(value=config['PATHS']['base_path'])
    # network_security_config_path_var = tk.StringVar(value=config['PATHS']['network_security_config_path'])
    # network_security_config_with_r_path_var = tk.StringVar(value=config['PATHS']['network_security_config_with_r_path'])


    def update_input_path():
        new_path = select_file_path()
        if new_path:
            input_var.set(new_path)

    def update_output_path():
        new_path = select_file_path()
        if new_path:
            output_var.set(new_path)

    def update_base_path():
        new_path = select_folder_path()
        if new_path:
            base_var.set(new_path)

    def update_network_security_config_path_path():
        new_path = select_file_path()
        if new_path:
            network_security_config_path_var.set(new_path)

    def update_network_security_config_with_r_path():
        new_path = select_file_path()
        if new_path:
            network_security_config_with_r_path_var.set(new_path)

    tk.Label(set_path_window, text="Excel 입력 파일 경로 :").grid(row=0, column=0, padx=10, pady=5)
    tk.Entry(set_path_window, textvariable=input_var, width=50).grid(row=0, column=1, padx=10, pady=5)
    tk.Button(set_path_window, text="변경", command=update_input_path).grid(row=0, column=2, padx=10, pady=5)

    tk.Label(set_path_window, text="Excel 출력 파일 경로 :").grid(row=1, column=0, padx=10, pady=5)
    tk.Entry(set_path_window, textvariable=output_var, width=50).grid(row=1, column=1, padx=10, pady=5)
    tk.Button(set_path_window, text="변경", command=update_output_path).grid(row=1, column=2, padx=10, pady=5)

    tk.Label(set_path_window, text="Base 폴더 경로 :").grid(row=2, column=0, padx=10, pady=5)
    tk.Entry(set_path_window, textvariable=base_var, width=50).grid(row=2, column=1, padx=10, pady=5)
    tk.Button(set_path_window, text="변경", command=update_base_path).grid(row=2, column=2, padx=10, pady=5)

    tk.Label(set_path_window, text="network_security_config 파일 경로 :").grid(row=3, column=0, padx=10, pady=5)
    tk.Entry(set_path_window, textvariable=network_security_config_path_var, width=50).grid(row=3, column=1, padx=10, pady=5)
    tk.Button(set_path_window, text="변경", command=update_network_security_config_path_path).grid(row=3, column=2, padx=10, pady=5)

    tk.Label(set_path_window, text="network_security_config_with_r 파일 경로 :").grid(row=4, column=0, padx=10, pady=5)
    tk.Entry(set_path_window, textvariable=network_security_config_with_r_path_var, width=50).grid(row=4, column=1, padx=10, pady=5)
    tk.Button(set_path_window, text="변경", command=update_network_security_config_with_r_path).grid(row=4, column=2, padx=10, pady=5)
    def save_and_close():
        if not input_var.get() or not output_var.get() or not base_var.get() or not network_security_config_path_var.get() or not network_security_config_with_r_path_var.get():
            messagebox.showwarning("경고", "모든 경로를 설정해야 합니다.")
            return  # 경로가 지정되지 않은 경우 함수를 종료하고 계속 실행     
           
        config['PATHS']['input_path'] = input_var.get()
        config['PATHS']['output_path'] = output_var.get()
        config['PATHS']['base_path'] = base_var.get()
        config['PATHS']['network_security_config_path'] = network_security_config_path_var.get()      
        config['PATHS']['network_security_config_with_r_path'] = network_security_config_with_r_path_var.get()        

        save_config(config)
        set_path_window.destroy()
        return
        # root.deiconify()

    tk.Button(set_path_window, text=" 저장 ", command=save_and_close).grid(row=5, column=0, columnspan=3, pady=10)

    set_path_window.mainloop()

    # return config['PATHS']['input_path'], config['PATHS']['output_path']
    return config


excel_input_path = None
excel_output_path = None
base_dir = None
network_security_config_path = None
network_security_config_with_r_path = None

def excelStartPoint(excel_input_path, excel_output_path, base_dir, network_security_config_path, network_security_config_with_r_path):


    # 엑셀 불러오기
    wb = openpyxl.load_workbook(excel_input_path)
    ws = wb['Sheet2']

    try:
        def confirm(root):
            nonlocal startPoint, endPoint
            try:
                startPoint = int(ent1.get())
                endPoint = int(ent2.get())
                if startPoint > endPoint:
                    messagebox.showerror("입력 오류", "시작점이 끝점보다 수가 적어야 합니다")
                    return  # 잘못된 입력이므로 확인 버튼을 다시 누르도록 함
                root.destroy()
            except ValueError:
                messagebox.showerror("입력 오류", "숫자를 입력해 주세요")


        
        def path_change(root):
            global excel_input_path, excel_output_path, base_dir, network_security_config_path, network_security_config_with_r_path
            root.withdraw()
            config = set_paths_gui(root)
            (excel_input_path,
            excel_output_path, 
            base_dir, 
            network_security_config_path, 
            network_security_config_with_r_path) = initialize_paths(config, root)
            root.deiconify()  # root 창 다시 표시

        startPoint = None
        endPoint = None
            
        root = Tk()
        root.title("앱 패키지 범위 지정")
        # root.attributes('-topmost', True)

        # 창을 화면 가운데로 이동
        center_window(root, 400, 300)
        def on_closing():
            root.destroy()
            sys.exit()

        root.protocol("WM_DELETE_WINDOW", on_closing)

        relative_path = os.path.join("images", "logo.png")

        lab_d = tk.Label(root)
        img = tk.PhotoImage(file = relative_path, master = root)
        lab_d.config(image = img)
        lab_d.pack()


        #startPoint 라벨
        lab1 = tk.Label(root)
        lab1.config(text = 'Excel 시작점')
        lab1.pack()

        #startPoint 입력창
        ent1 = tk.Entry(root)

        ent1.pack()


        #endPoint 라벨
        lab2 = tk.Label(root)
        lab2.config(text = 'Excel 끝점')
        lab2.pack()

        #endPoint 입력창
        ent2 = tk.Entry(root)

        ent2.pack()

        #확인

        btn = tk.Button(root)
        btn.config(text = '확인 및 실행')

        btn.config(command=lambda: confirm(root))
        btn.pack()

        #파일 경로 변경

        btn = tk.Button(root)
        btn.config(text = '파일 경로 변경')

        btn.config(command=lambda: path_change(root))
        btn.pack()
                
        root.mainloop()
    except Exception as e:
        return e

    return startPoint, endPoint, ws, excel_input_path, excel_output_path, base_dir, network_security_config_path, network_security_config_with_r_path

def excelEndPoint(output_path, ranking, category, app_name, package_name, totaluser, totaltime, monthuser, result, error, detail):


    if not output_path:
        raise ValueError("저장할 파일 경로를 선택해 주세요.")

    wb = openpyxl.load_workbook(output_path)
    ws = wb['Sheet1']

    last_row = ws.max_row + 1  # 다음에 기록할 행 번호

    ws[f'A{last_row}'] = ranking
    ws[f'B{last_row}'] = category
    ws[f'C{last_row}'] = app_name
    ws[f'D{last_row}'] = package_name
    ws[f'E{last_row}'] = totaluser
    ws[f'F{last_row}'] = totaltime
    ws[f'G{last_row}'] = monthuser
    ws[f'H{last_row}'] = result
    ws[f'I{last_row}'] = error
    ws[f'J{last_row}'] = detail




    wb.save(output_path)
